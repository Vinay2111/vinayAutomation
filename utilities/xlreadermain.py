import openpyxl


def getRowCount(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColCount(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column


def getCellData(path,sheetName,rowNum,colNum):     #to read the data
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum,column=colNum).value

def setCellData(path,sheetName,rowNum,colNum,data):     #to write the data
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(path)

path="..//excel//testdata.xlsx"
sheetName="login"
rows=getRowCount(path,sheetName)
cols=getColCount(path,sheetName)

print(rows,"--",cols)

print(getCellData(path,sheetName,1,2))
