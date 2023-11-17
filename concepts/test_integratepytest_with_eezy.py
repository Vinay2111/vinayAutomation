import time

import openpyxl
import pytest
from appium import webdriver
from appium.webdriver.appium_service import AppiumService
from appium.webdriver.common.appiumby import AppiumBy
from utilities import dataprovider

def get_data():
    '''return [
        ["qwerty@ymail.com","111111"],
        ["test@test.test","qwerty123"]
    ]'''
    workbook=openpyxl.load_workbook("..//excel//testdata.xlsx")
    sheet=workbook["login"]
    totalrows=sheet.max_row
    totalcols=sheet.max_colomn
    mainList=[]

    for i in range(2,totalrows+1):
        dataList=[]
        for j in range(1,totalcols+1):
            data=sheet.cell(row=i,column=j).value
            dataList.insert(j,data)
        mainList.insert(i,dataList)
        return mainList

def setup_function():
    global appium_service
    #appium_service=AppiumService()
   # appium_service.start()

desired_cap= {
    "appium:deviceName": "6a45d0657d05",
    "platformName": "Android",
    "appium:appPackage": "com.eezy.ai.dev",
    "appium:appWaitActivity": "com.natife.eezy.MainActivity",
    "appium:app": "/Users/vinay/Documents/appium/Android APK/app-dev-debug.apk",
}
global driver
driver=webdriver.Remote("http://localhost:4723/wd/hub",desired_cap)
driver.implicitly_wait(30)

def teardown_function():
    time.sleep(2)
    #driver.quit()
    #appium_service.stop()


@pytest.mark.parametrize("email,password",dataprovider.get_data("login"))
def test_login(email,password):
    driver.find_element(AppiumBy.ID, 'com.eezy.ai.dev:id/negativeButton').click()
    driver.find_element(AppiumBy.ID, 'com.eezy.ai.dev:id/wizardBottomButton').click()
    driver.find_element(AppiumBy.ID, 'com.eezy.ai.dev:id/moreLoginOptions').click()
    driver.find_element(AppiumBy.ID, 'com.eezy.ai.dev:id/emailSignUp').click()
    driver.implicitly_wait(5)

    driver.find_element(AppiumBy.ID, 'com.eezy.ai.dev:id/signUpAlreadyHaveAccountTextView').click()
    driver.find_element(AppiumBy.ID, 'com.eezy.ai.dev:id/moreLoginOptions').click()
    driver.find_element(AppiumBy.ID, 'com.eezy.ai.dev:id/emailSignUp').click()
    driver.implicitly_wait(4)
    driver.find_element(AppiumBy.ID, 'com.eezy.ai.dev:id/loginEmailEditText').send_keys(email)
    driver.find_element(AppiumBy.ID, 'com.eezy.ai.dev:id/loginPasswordEditText').send_keys(password)
    driver.implicitly_wait(3)
    driver.find_element(AppiumBy.ID, 'com.eezy.ai.dev:id/loginNowButton').click()

    print(email,password)

